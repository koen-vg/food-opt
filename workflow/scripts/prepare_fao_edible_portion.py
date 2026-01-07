# SPDX-FileCopyrightText: 2025 Koen van Greevenbroek
#
# SPDX-License-Identifier: GPL-3.0-or-later

"""Extract edible portion coefficients from the FAO nutrient table.

The script reads sheet "03" of the FAO Nutrient Conversion Table workbook and
pulls the *edible portion coefficient* for each FAOSTAT item mapped to the
model's configured crops. The resulting CSV lists the coefficient alongside the
FAOSTAT item code and edible portion type flag supplied by FAO.

Special handling is applied to certain crops where FAO's coefficient reflects
processed products rather than whole-crop dry matter:
- Grains (rice, barley, oat, buckwheat): FAO gives milled/hulled conversion;
  we force to 1.0 for whole grain, handling milling separately.
- Sugar crops (sugarcane, sugarbeet) and oil-palm: we operate on whole-crop dry
  matter, so the edible portion is forced to 1.0 and downstream processing is
  handled in ``data/foods.csv``.
"""

import csv
from dataclasses import dataclass
import logging
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from workflow.scripts.logging_config import setup_script_logging

# Logger will be configured in __main__ block
logger = logging.getLogger(__name__)

ALTERNATE_ITEM_NAMES: dict[str, str] = {
    "rape or colza seed": "rapeseed or colza seed",
    "oil palm fruit": "palm-oil",
}

# Crops for which the edible portion coefficient should be set to 1.0 despite
# FAO listing a lower value. Reasons vary by crop type:
# - Grains (rice, barley, oat, buckwheat): FAO coefficient represents milled/
#   hulled grain; we track whole grain and handle milling separately.
# - Sugar crops (sugarcane, sugarbeet) and oil-palm: the model converts
#   reported yields back to whole-crop dry matter prior to processing.
EDIBLE_PORTION_EXCEPTIONS: set[str] = {
    "dryland-rice",
    "wetland-rice",
    "barley",
    "oat",
    "buckwheat",
    "oil-palm",
    "sugarcane",
    "sugarbeet",
}

FALLBACK_FULL_EDIBLE: set[str] = {"oil-palm", "sugarcane", "sugarbeet"}


@dataclass
class ComponentRow:
    code: str
    description: str
    edible_coefficient: Optional[float]
    edible_type: Optional[int]


def _coerce_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        logger.warning("Could not parse edible coefficient value %r", value)
        return None


def _coerce_int(value) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        logger.warning("Could not parse edible portion type value %r", value)
        return None


def _load_component_values(xlsx_path: Path) -> dict[str, ComponentRow]:
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["03"]
    except KeyError as exc:
        raise ValueError("Worksheet '03' with component values not found") from exc

    records: dict[str, ComponentRow] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = row[0]
        description = row[1]
        if not code or not description:
            continue

        edible_coefficient = _coerce_float(row[4])
        edible_type = _coerce_int(row[5])

        key = str(description).strip().lower()
        records[key] = ComponentRow(
            code=str(code).strip(),
            description=str(description).strip(),
            edible_coefficient=edible_coefficient,
            edible_type=edible_type,
        )

    return records


def _read_crop_mapping(path: Path) -> dict[str, Optional[str]]:
    mapping: dict[str, Optional[str]] = {}
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            crop = row.get("crop")
            item = row.get("faostat_item")
            if crop is None:
                continue
            item = item.strip() if item else None
            mapping[crop] = item if item else None
    return mapping


def main() -> None:
    crops: list[str] = list(snakemake.params.crops)  # type: ignore[name-defined]
    xlsx_path = Path(snakemake.input.table)  # type: ignore[name-defined]
    mapping_path = Path(snakemake.input.mapping)  # type: ignore[name-defined]
    output_path = Path(snakemake.output.edible_portion)  # type: ignore[name-defined]

    records_by_name = _load_component_values(xlsx_path)
    crop_to_item = _read_crop_mapping(mapping_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    missing_items: list[str] = []
    missing_components: list[str] = []

    with output_path.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "crop",
                "faostat_item",
                "fao_code",
                "edible_portion_coefficient",
                "edible_portion_type",
            ],
        )
        writer.writeheader()

        for crop in crops:
            item = crop_to_item.get(crop)
            if item is None:
                missing_items.append(crop)
                writer.writerow(
                    {
                        "crop": crop,
                        "faostat_item": "",
                        "fao_code": "",
                        "edible_portion_coefficient": "",
                        "edible_portion_type": "",
                    }
                )
                continue

            key = item.strip().lower()
            record = records_by_name.get(key)
            if record is None and key in ALTERNATE_ITEM_NAMES:
                record = records_by_name.get(ALTERNATE_ITEM_NAMES[key])
            if record is None:
                if crop in FALLBACK_FULL_EDIBLE:
                    writer.writerow(
                        {
                            "crop": crop,
                            "faostat_item": item,
                            "fao_code": "",
                            "edible_portion_coefficient": 1.0,
                            "edible_portion_type": "",
                        }
                    )
                    continue
                missing_components.append(item)
                writer.writerow(
                    {
                        "crop": crop,
                        "faostat_item": item,
                        "fao_code": "",
                        "edible_portion_coefficient": "",
                        "edible_portion_type": "",
                    }
                )
                continue

            # Apply exception: force edible_coefficient to 1.0 for certain crops
            edible_coeff = (
                1.0
                if crop in EDIBLE_PORTION_EXCEPTIONS
                else (
                    ""
                    if record.edible_coefficient is None
                    else record.edible_coefficient
                )
            )

            writer.writerow(
                {
                    "crop": crop,
                    "faostat_item": record.description,
                    "fao_code": record.code,
                    "edible_portion_coefficient": edible_coeff,
                    "edible_portion_type": (
                        "" if record.edible_type is None else record.edible_type
                    ),
                }
            )

    if missing_items:
        logger.warning(
            "No FAOSTAT item mapping for crops: %s", ", ".join(sorted(missing_items))
        )
    if missing_components:
        logger.warning(
            "No component entry found in FAO table for items: %s",
            ", ".join(sorted(set(missing_components))),
        )


if __name__ == "__main__":
    # Configure logging
    logger = setup_script_logging(log_file=snakemake.log[0] if snakemake.log else None)

    main()
